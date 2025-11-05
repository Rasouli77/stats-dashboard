from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from datetime import datetime, timedelta
import uuid
import openpyxl.styles
from .models import (
    PeopleCounting,
    Branch,
    UserProfile,
    PermissionToViewBranch,
    Campaign,
    Cam,
    Invoice,
    AlertCameraMalfunction,
    AlertCameraMalfunctionMessage,
)
from django.db.models import Sum, F, Q, Min, Max
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from .forms import (
    Generate_User,
    UserPermissions,
    AssignBranchPermissions,
    CreateCampaign,
    UploadInvoiceExcel,
    InvoiceForm,
    AlertCameraMalfunctionForm,
)
import jdatetime
import json
from django.contrib import messages
from django.db.models import F
from .camera_data import get_custom_date_camera_data, update_or_create_camera_data
from django.db import connection
from django.utils import timezone
import subprocess
import openpyxl
from io import BytesIO
import random
import requests
from django.db.models import F, ExpressionWrapper, FloatField


def perm_to_open(request, url_hash):
    """Checks if the user has the right to open the page."""
    if not request.user.is_active:
        return False
    merchant = request.user.profile.merchant
    try:
        if merchant.url_hash:
            if merchant.url_hash != url_hash:
                return False
            else:
                return True
        else:
            return False
    except Exception as e:
        return False


def jalali_to_gregorian(date_str: str):
    """Convert a Jalali date (datetime.date or datetime.datetime) to Gregorian (YYYY-MM-DD)."""
    try:
        jyear, jmonth, jday = map(int, date_str.split("-"))
        gregorian_date = jdatetime.date(jyear, jmonth, jday).togregorian()
        return gregorian_date
    except Exception as e:
        return None


def convert_gregorian_to_jalali(g_date):
    """Convert a Gregorian date (datetime.date or datetime.datetime) to Jalali (YYYY-MM-DD)."""
    if not g_date:
        return ""
    j_date = jdatetime.date.fromgregorian(date=g_date)
    return j_date.strftime("%Y-%m-%d")


@login_required
def people_counter(request, url_hash):
    """This provides an overview for people counters in all branches."""
    # Rights
    if not perm_to_open(request, url_hash):
        return render(request, "401.html", status=401)
    queryset = []
    if request.user.profile.is_manager:
        branches = Branch.objects.only("name", "pk").filter(merchant__url_hash=url_hash)
    else:
        permitted_branches = (
            PermissionToViewBranch.objects.defer("date_created", "last_modified")
            .select_related("branch")
            .filter(user__pk=request.user.profile.pk)
        )
        permitted_branches_list = []
        for permitted_branch in permitted_branches:
            permitted_branches_list.append(permitted_branch.branch.pk)
        branches = Branch.objects.only("name", "pk").filter(
            merchant__url_hash=url_hash, pk__in=permitted_branches_list
        )
    selected_branches_str = request.GET.getlist("branch")
    start_date_str = str(jalali_to_gregorian(request.GET.get("start-date")))
    end_date_str = str(jalali_to_gregorian(request.GET.get("end-date")))
    start_date = 0
    end_date = 0
    entry_totals = []
    exit_totals = []
    campaign_list = []
    selected_branches = []
    branches_stats = {}
    if start_date_str != "None" and end_date_str != "None":
        try:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
        except Exception as e:
            print(e)
        queryset = (
            PeopleCounting.objects.defer("date_created", "last_modified", "cam")
            .filter(merchant__url_hash=url_hash, date__range=(start_date, end_date))
            .values("date")
            .annotate(total_entry=Sum("entry"))
            .order_by("date")
        )
        permissions = (
            PermissionToViewBranch.objects.defer("date_created", "last_modified")
            .select_related("branch")
            .filter(user__pk=request.user.profile.pk)
        )
        if len(permissions) == 0 and not request.user.profile.is_manager:
            return render(request, "401.html", status=401)
        if request.user.profile.is_manager:
            branches = Branch.objects.only("name", "pk").filter(
                merchant__url_hash=url_hash
            )
            campaigns = (
                Campaign.objects.filter(
                    branch__merchant__url_hash=request.user.profile.merchant.url_hash
                )
                .filter(Q(start_date__lte=end_date) & Q(end_date__gte=start_date))
                .values("group_id")
                .annotate(
                    campaign_name=Min("name"),
                    campaign_start_date=Min("start_date"),
                    campaign_end_date=Max("end_date"),
                    campaign_cost=Sum("cost"),
                    campaign_type=Min("campaign_type"),
                    campaign_last_modified=Max("last_modified"),
                    campaign_group_id=Min("group_id"),
                )
                .order_by("-campaign_last_modified")
            )
        else:
            permitted_branches = (
                PermissionToViewBranch.objects.defer("date_created", "last_modified")
                .select_related("branch")
                .filter(user__pk=request.user.profile.pk)
            )
            permitted_branches_list = []
            for permitted_branch in permitted_branches:
                permitted_branches_list.append(permitted_branch.branch.pk)
            branches = Branch.objects.only("name", "pk").filter(
                merchant__url_hash=url_hash, pk__in=permitted_branches_list
            )
            campaigns = (
                Campaign.objects.filter(
                    branch__merchant__url_hash=request.user.profile.merchant.url_hash
                )
                .filter(branch__pk__in=permitted_branches_list)
                .filter(Q(start_date__lte=end_date) & Q(end_date__gte=start_date))
                .values("group_id")
                .annotate(
                    campaign_name=Min("name"),
                    campaign_start_date=Min("start_date"),
                    campaign_end_date=Max("end_date"),
                    campaign_cost=Sum("cost"),
                    campaign_type=Min("campaign_type"),
                    campaign_last_modified=Max("last_modified"),
                    campaign_group_id=Min("group_id"),
                )
                .order_by("-campaign_last_modified")
            )
            campaign_list = []
            dictionary = {}
            for campaign in campaigns:
                dictionary = {
                    "campaign_name": campaign["campaign_name"],
                    "campaign_start_date": campaign["campaign_start_date"],
                    "campaign_end_date": campaign["campaign_end_date"],
                    "campaign_cost": campaign["campaign_cost"],
                    "campaign_type": campaign["campaign_type"],
                }
                campaign_list.append(dictionary)
            queryset = queryset.filter(branch__pk__in=permitted_branches_list)
    if selected_branches_str:
        selected_branches = [int(pk) for pk in selected_branches_str]
        try:
            queryset = queryset.filter(branch__pk__in=selected_branches)
            if request.user.profile.is_manager:
                campaigns = (
                    Campaign.objects.filter(
                        branch__merchant__url_hash=request.user.profile.merchant.url_hash
                    )
                    .filter(branch__pk__in=selected_branches)
                    .filter(Q(start_date__lte=end_date) & Q(end_date__gte=start_date))
                    .values("group_id")
                    .annotate(
                        campaign_name=Min("name"),
                        campaign_start_date=Min("start_date"),
                        campaign_end_date=Max("end_date"),
                        campaign_cost=Sum("cost"),
                        campaign_type=Min("campaign_type"),
                        campaign_last_modified=Max("last_modified"),
                        campaign_group_id=Min("group_id"),
                    )
                    .order_by("-campaign_last_modified")
                )
            else:
                permitted_branches = (
                    PermissionToViewBranch.objects.defer(
                        "date_created", "last_modified"
                    )
                    .select_related("branch")
                    .filter(user__pk=request.user.profile.pk)
                )
                permitted_branches_list = []
                for permitted_branch in permitted_branches:
                    permitted_branches_list.append(permitted_branch.branch.pk)
                branches = Branch.objects.only("name", "pk").filter(
                    merchant__url_hash=url_hash, pk__in=permitted_branches_list
                )
                campaigns = (
                    Campaign.objects.filter(
                        branch__merchant__url_hash=request.user.profile.merchant.url_hash
                    )
                    .filter(branch__pk__in=permitted_branches_list)
                    .filter(Q(start_date__lte=end_date) & Q(end_date__gte=start_date))
                    .values("group_id")
                    .annotate(
                        campaign_name=Min("name"),
                        campaign_start_date=Min("start_date"),
                        campaign_end_date=Max("end_date"),
                        campaign_cost=Sum("cost"),
                        campaign_type=Min("campaign_type"),
                        campaign_last_modified=Max("last_modified"),
                        campaign_group_id=Min("group_id"),
                    )
                    .order_by("-campaign_last_modified")
                )
            campaign_list = []
            dictionary = {}
            for campaign in campaigns:
                dictionary = {
                    "campaign_name": campaign["campaign_name"],
                    "campaign_start_date": campaign["campaign_start_date"],
                    "campaign_end_date": campaign["campaign_end_date"],
                    "campaign_cost": campaign["campaign_cost"],
                    "campaign_type": campaign["campaign_type"],
                }
                campaign_list.append(dictionary)
            entry_totals = [float(row["total_entry"]) for row in queryset]
            exit_totals = []

        except Exception as e:
            print(e)

    entry_totals = [float(row["total_entry"]) for row in queryset]
    exit_totals = []
    dates = [str(row["date"].strftime("%Y-%m-%d")) for row in queryset]

    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        print(campaign_list)
        return JsonResponse(
            {
                "dates": dates,
                "entry_totals": entry_totals,
                "exit_totals": exit_totals,
                "campaigns": campaign_list,
            }
        )
    print(connection.queries, len(connection.queries))
    return render(
        request,
        "people-counter.html",
        {
            "dates": json.dumps(dates),
            "entry_totals": json.dumps(entry_totals),
            "exit_totals": json.dumps(exit_totals),
            "branches": branches,
            "branches_data": json.dumps(dict(branches_stats)),
            "campaigns": json.dumps(campaign_list),
        },
    )


@login_required
def users_list(request, url_hash):
    """A list of all users for managers."""
    # Rights
    if not perm_to_open(request, url_hash):
        return render(request, "401.html", status=401)
    users = UserProfile.objects.filter(merchant__url_hash=url_hash)
    if (
        request.user.profile.merchant.url_hash == url_hash
        and request.user.profile.is_manager
        and request.user.is_active
    ):
        return render(request, "users.html", {"users": users})
    return render(request, "401.html", status=401)


@login_required
def generate_user(request, url_hash):
    """Using this, managers can create users."""
    # Rights
    if not perm_to_open(request, url_hash):
        return render(request, "401.html", status=401)
    if (
        request.user.profile.merchant.url_hash == url_hash
        and request.user.profile.is_manager
        and request.user.is_active
    ):
        if request.method == "POST":
            form = Generate_User(request.POST)
            if form.is_valid():
                new_user = form.save()
                try:
                    UserProfile.objects.create(
                        user=new_user,
                        merchant=request.user.profile.merchant,
                        mobile="",
                        is_manager=False,
                    )
                    messages.success(request, "کاربر با موفقیت ساخته شد")
                except Exception as e:
                    print(e)
                    messages.error(request, f"{e}")
        else:
            form = Generate_User()
        return render(request, "create-user.html", {"form": form})
    return render(request, "401.html", status=401)


@login_required
def user_permissions(request, user_id):
    user = get_object_or_404(User, pk=user_id)
    if (
        request.user.profile.merchant.url_hash == user.profile.merchant.url_hash
        and request.user.is_active
        and request.user.profile.is_manager
    ):
        if request.method == "POST":
            form = UserPermissions(request.POST, instance=user)
            if form.is_valid():
                form.save()
                return redirect("profile", user.pk)
        else:
            form = UserPermissions(instance=user)
        return render(request, "user-permissions.html", {"user": user, "form": form})
    return render(request, "401.html", status=401)


@login_required
def branch_permissions(request, user_id):
    user_profile = UserProfile.objects.get(user=user_id)
    perms = (
        PermissionToViewBranch.objects.defer("date_created", "last_modified")
        .select_related("branch")
        .filter(user=user_profile)
    )
    branches = Branch.objects.defer(
        "country", "province", "city", "district", "date_created", "last_modified"
    ).filter(merchant=user_profile.merchant)
    branch_count = branches.count()
    allowed_branches = []
    for perm in perms:
        allowed_branches.append(perm.branch.pk)
    if len(allowed_branches) > 0:
        branches = branches.exclude(pk__in=allowed_branches)
    if (
        request.user.is_active
        and request.user.profile.is_manager
        and request.user.profile.merchant.url_hash == user_profile.merchant.url_hash
    ):
        if request.method == "POST":
            form = AssignBranchPermissions(request.POST)
            if form.is_valid():
                form.save()
                return redirect("edit-branch-permissions", user_profile.user.pk)
        else:
            form = AssignBranchPermissions()
        return render(
            request,
            "branch-permissions.html",
            {
                "form": form,
                "user_profile_id": user_profile.pk,
                "perms": perms,
                "branches": branches,
                "branch_count": branch_count,
                "allowed_branch_count": len(allowed_branches),
            },
        )
    return render(request, "401.html", status=401)


@login_required
def home(request, url_hash):
    """A general overview of certain aggregations"""
    # Rights
    if not perm_to_open(request, url_hash):
        return render(request, "401.html", status=401)
    # people counting queryset
    queryset = PeopleCounting.objects.defer(
        "date_created", "last_modified", "exit"
    ).filter(merchant__url_hash=url_hash)
    # invoice queryset
    invoice_queryset = Invoice.objects.defer("date_created", "last_modified").filter(
        branch__merchant__url_hash=url_hash
    )
    # 7 days
    today = timezone.now().today()
    last_7_days_start = today - timedelta(days=8)
    last_7_days_end = today - timedelta(days=1)
    previous_7_days_start = today - timedelta(days=15)
    previous_7_days_end = today - timedelta(days=8)

    # entries
    last_7_days_entry = queryset.filter(
        date__range=(last_7_days_start, last_7_days_end)
    ).aggregate(total_entry=Sum("entry"))
    previous_7_days_entry = queryset.filter(
        date__range=(previous_7_days_start, previous_7_days_end)
    ).aggregate(total_entry=Sum("entry"))
    # invoice number
    last_7_days_invoice_number = invoice_queryset.filter(
        date__range=(last_7_days_start, last_7_days_end)
    ).aggregate(total_number=Sum("total_items"))
    previous_7_days_invoice_number = invoice_queryset.filter(
        date__range=(previous_7_days_start, previous_7_days_end)
    ).aggregate(total_number=Sum("total_items"))
    # invoice amount
    last_7_days_invoice_amount = invoice_queryset.filter(
        date__range=(last_7_days_start, last_7_days_end)
    ).aggregate(total_amount=Sum("total_amount"))
    previous_7_days_invoice_amount = invoice_queryset.filter(
        date__range=(previous_7_days_start, previous_7_days_end)
    ).aggregate(total_amount=Sum("total_amount"))
    # invoice product count
    last_7_days_invoice_product = invoice_queryset.filter(
        date__range=(last_7_days_start, last_7_days_end)
    ).aggregate(total_product=Sum("total_product"))
    previous_7_days_invoice_product = invoice_queryset.filter(
        date__range=(previous_7_days_start, previous_7_days_end)
    ).aggregate(total_product=Sum("total_product"))
    try:
        # people counting difference percentage
        diff_7_per = (
            (last_7_days_entry["total_entry"] - previous_7_days_entry["total_entry"])
            / previous_7_days_entry["total_entry"]
        ) * 100
        rounded_diff_7_per = round(diff_7_per, 2)
    except:
        diff_7_per = 0
        rounded_diff_7_per = 0

    try:
        # invoice number difference percentage
        diff_7_invoice_num = (
            (
                last_7_days_invoice_number["total_number"]
                - previous_7_days_invoice_number["total_number"]
            )
            / previous_7_days_invoice_number["total_number"]
        ) * 100
        rounded_diff_7_invoice_num = round(diff_7_invoice_num, 2)
    except:
        diff_7_invoice_num = 0
        rounded_diff_7_invoice_num = 0

    try:
        # invoice amount difference percentage
        diff_7_invoice_amount = (
            (
                last_7_days_invoice_amount["total_amount"]
                - previous_7_days_invoice_amount["total_amount"]
            )
            / previous_7_days_invoice_amount["total_amount"]
        ) * 100
        rounded_diff_7_invoice_amount = round(diff_7_invoice_amount, 2)
    except:
        diff_7_invoice_amount = 0
        rounded_diff_7_invoice_amount = 0

    try:
        # invoice product count difference
        diff_7_invoice_product = (
            (
                last_7_days_invoice_product["total_product"]
                - previous_7_days_invoice_product["total_product"]
            )
            / previous_7_days_invoice_product["total_product"]
        ) * 100
        rounded_diff_7_invoice_product = round(diff_7_invoice_product, 2)
    except:
        diff_7_invoice_product = 0
        rounded_diff_7_invoice_product = 0

    # branches
    last_7_days_best_branch_name = (
        queryset.filter(date__range=(last_7_days_start, last_7_days_end))
        .values("branch__id", "branch__name")
        .annotate(total_entry=Sum("entry"))
        .order_by("-total_entry")
        .first()
    )
    last_7_days_worst_branch_name = (
        queryset.filter(date__range=(last_7_days_start, last_7_days_end))
        .values("branch__id", "branch__name")
        .annotate(total_entry=Sum("entry"))
        .order_by("total_entry")
        .first()
    )

    branch_7_days_traffic_share = []
    # branch rank
    try:
        branch_7_day_ranks = (
            queryset.filter(date__range=(last_7_days_start, last_7_days_end))
            .values("branch__id", "branch__name")
            .annotate(total_entry=Sum("entry"))
            .order_by("-total_entry")
        )
        for item in branch_7_day_ranks:
            branch_7_days_traffic_share.append(
                {"name": item["branch__name"], "y": float(item["total_entry"])}
            )
    except:
        branch_7_day_ranks = []

    # 30 days
    last_30_days_start = today - timedelta(days=30)
    last_30_days_end = today
    previous_30_days_start = today - timedelta(days=60)
    previous_30_days_end = today - timedelta(days=30)

    # entries
    last_30_days_entry = queryset.filter(
        date__range=(last_30_days_start, last_30_days_end)
    ).aggregate(total_entry=Sum("entry"))
    previous_30_days_entry = queryset.filter(
        date__range=(previous_30_days_start, previous_30_days_end)
    ).aggregate(total_entry=Sum("entry"))
    # invoice number
    last_30_days_invoice_number = invoice_queryset.filter(
        date__range=(last_30_days_start, last_30_days_end)
    ).aggregate(total_number=Sum("total_items"))
    previous_30_days_invoice_number = invoice_queryset.filter(
        date__range=(previous_30_days_start, previous_30_days_end)
    ).aggregate(total_number=Sum("total_items"))
    # invoice amount
    last_30_days_invoice_amount = invoice_queryset.filter(
        date__range=(last_30_days_start, last_30_days_end)
    ).aggregate(total_amount=Sum("total_amount"))
    previous_30_days_invoice_amount = invoice_queryset.filter(
        date__range=(previous_30_days_start, previous_30_days_end)
    ).aggregate(total_amount=Sum("total_amount"))
    # invoice product count
    last_30_days_invoice_product = invoice_queryset.filter(
        date__range=(last_30_days_start, last_30_days_end)
    ).aggregate(total_product=Sum("total_product"))
    previous_30_days_invoice_product = invoice_queryset.filter(
        date__range=(previous_30_days_start, previous_30_days_end)
    ).aggregate(total_product=Sum("total_product"))
    try:
        diff_30_per = (
            (last_30_days_entry["total_entry"] - previous_30_days_entry["total_entry"])
            / previous_30_days_entry["total_entry"]
        ) * 100
        rounded_diff_30_per = round(diff_30_per, 2)
    except:
        diff_30_per = 0
        rounded_diff_30_per = 0

    try:
        # invoice number difference percentage
        diff_30_invoice_num = (
            (
                last_30_days_invoice_number["total_number"]
                - previous_30_days_invoice_number["total_number"]
            )
            / previous_30_days_invoice_number["total_number"]
        ) * 100
        rounded_diff_30_invoice_num = round(diff_30_invoice_num, 2)
    except:
        diff_30_invoice_num = 0
        rounded_diff_30_invoice_num = 0

    try:
        # invoice amount difference percentage
        diff_30_invoice_amount = (
            (
                last_30_days_invoice_amount["total_amount"]
                - previous_30_days_invoice_amount["total_amount"]
            )
            / previous_30_days_invoice_amount["total_amount"]
        ) * 100
        rounded_diff_30_invoice_amount = round(diff_30_invoice_amount, 2)
    except:
        diff_30_invoice_amount = 0
        rounded_diff_30_invoice_amount = 0

    try:
        # invoice product count difference
        diff_30_invoice_product = (
            (
                last_30_days_invoice_product["total_product"]
                - previous_30_days_invoice_product["total_product"]
            )
            / previous_30_days_invoice_product["total_product"]
        ) * 100
        rounded_diff_30_invoice_product = round(diff_30_invoice_product, 2)
    except:
        diff_30_invoice_product = 0
        rounded_diff_30_invoice_product = 0

    # branches
    last_30_days_best_branch_name = (
        queryset.filter(date__range=(last_30_days_start, last_30_days_end))
        .values("branch__id", "branch__name")
        .annotate(total_entry=Sum("entry"))
        .order_by("-total_entry")
        .first()
    )
    last_30_days_worst_branch_name = (
        queryset.filter(date__range=(last_30_days_start, last_30_days_end))
        .values("branch__id", "branch__name")
        .annotate(total_entry=Sum("entry"))
        .order_by("total_entry")
        .first()
    )
    branch_30_days_traffic_share = []
    # branch rank
    try:
        branch_30_day_ranks = (
            queryset.filter(date__range=(last_30_days_start, last_30_days_end))
            .values("branch__id", "branch__name")
            .annotate(total_entry=Sum("entry"))
            .order_by("-total_entry")
        )
        for item in branch_30_day_ranks:
            branch_30_days_traffic_share.append(
                {"name": item["branch__name"], "y": float(item["total_entry"])}
            )
    except:
        branch_30_day_ranks = []
    # online data
    today = datetime.today()
    today_data = []
    try:
        branches_traffic_today = queryset.filter(date=today)
        for item in branches_traffic_today:
            today_data_item = []
            today_data_item.append(float(item.cam.google_longitude))
            today_data_item.append(float(item.cam.google_latitude))
            today_data_item.append(float(item.entry))
            flag = True
            for i in today_data:
                if today_data_item[0] == i[0] and today_data_item[1] == i[1]:
                    flag = False
                    i[2] += today_data_item[2]
            if flag:
                today_data.append(today_data_item)
    except:
        pass
    print(branch_7_days_traffic_share)
    print(branch_30_days_traffic_share)
    return render(
        request,
        "home.html",
        {
            "last_7_days_best_branch_name": last_7_days_best_branch_name,
            "last_7_days_worst_branch_name": last_7_days_worst_branch_name,
            "branch_30_days_traffic_share": json.dumps(branch_30_days_traffic_share),
            "branch_7_days_traffic_share": json.dumps(branch_7_days_traffic_share),
            "last_30_days_best_branch_name": last_30_days_best_branch_name,
            "last_30_days_worst_branch_name": last_30_days_worst_branch_name,
            "rounded_diff_30_per": rounded_diff_30_per,
            "rounded_diff_7_per": rounded_diff_7_per,
            "last_30_days_entry_total": last_30_days_entry["total_entry"],
            "last_7_days_entry_total": last_7_days_entry["total_entry"],
            "previous_7_days_entry_total": previous_7_days_entry["total_entry"],
            "previous_30_days_entry_total": previous_30_days_entry["total_entry"],
            "online_map_data": json.dumps(today_data),
            "last_7_days_invoice_number": last_7_days_invoice_number["total_number"],
            "last_7_days_invoice_amount": last_7_days_invoice_amount["total_amount"]
            / 10,
            "last_7_days_invoice_product": last_7_days_invoice_product["total_product"],
            "rounded_diff_7_invoice_num": rounded_diff_7_invoice_num,
            "rounded_diff_7_invoice_amount": rounded_diff_7_invoice_amount,
            "rounded_diff_7_invoice_product": rounded_diff_7_invoice_product,
            "last_30_days_invoice_number": last_30_days_invoice_number["total_number"],
            "last_30_days_invoice_amount": last_30_days_invoice_amount["total_amount"]
            / 10,
            "last_30_days_invoice_product": last_30_days_invoice_product[
                "total_product"
            ],
            "rounded_diff_30_invoice_num": rounded_diff_30_invoice_num,
            "rounded_diff_30_invoice_amount": rounded_diff_30_invoice_amount,
            "rounded_diff_30_invoice_product": rounded_diff_30_invoice_product,
        },
    )


@login_required
def profile(request, user_id):
    """This loads a profile for each user."""
    user = get_object_or_404(User, pk=user_id)
    user_permissions = user.user_permissions.all()
    user_profile = UserProfile.objects.get(user=user.pk)
    allowed_branches = PermissionToViewBranch.objects.defer(
        "date_created", "last_modified"
    ).filter(user=user_profile.pk)
    if request.user.pk == user.pk and request.user.is_active:
        return render(
            request,
            "profile.html",
            {
                "user": user,
                "permissions": user_permissions,
                "branches": allowed_branches,
            },
        )

    if (
        request.user.profile.merchant.url_hash == user.profile.merchant.url_hash
        and request.user.profile.is_manager
        and request.user.is_active
    ):
        return render(
            request,
            "profile.html",
            {
                "user": user,
                "permissions": user_permissions,
                "branches": allowed_branches,
                "branch_length": len(allowed_branches),
            },
        )
    return render(request, "401.html", status=401)


@login_required
def campaign(request, url_hash):
    """A full list of campaigns if allowed"""
    # Rights
    if not perm_to_open(request, url_hash):
        return render(request, "401.html", status=401)
    if not request.user.profile.is_manager:
        return render(request, "401.html", status=401)
    campaigns = (
        Campaign.objects.filter(
            branch__merchant__url_hash=request.user.profile.merchant.url_hash
        )
        .values("group_id")
        .annotate(
            campaign_name=Min("name"),
            campaign_start_date=Min("start_date"),
            campaign_end_date=Max("end_date"),
            campaign_cost=Sum("cost"),
            campaign_type=Min("campaign_type"),
            campaign_last_modified=Max("last_modified"),
            campaign_group_id=Min("group_id"),
        )
        .order_by("-campaign_last_modified")
    )

    for campaign in campaigns:
        branch_names = (
            Campaign.objects.filter(group_id=campaign["campaign_group_id"])
            .values_list("branch__name", flat=True)
            .distinct()
        )
        campaign["branch_names"] = ", ".join(branch_names)
    return render(request, "campaign.html", {"campaigns": campaigns})


@login_required
def create_campaign(request, url_hash):
    """Using this, users can create campaigns."""
    # Rights
    if not perm_to_open(request, url_hash):
        return render(request, "401.html", status=401)
    if request.user.profile.merchant.url_hash == url_hash:
        permissions = (
            PermissionToViewBranch.objects.defer("date_created", "last_modified")
            .select_related("branch")
            .filter(user__pk=request.user.profile.pk)
        )
        permission_list = []
        for permission in permissions:
            permission_list.append(permission.branch.pk)
        branches = Branch.objects.defer(
            "country", "province", "city", "district", "date_created", "last_modified"
        ).filter(merchant__url_hash=url_hash, pk__in=permission_list)
        if request.user.profile.is_manager and request.user.is_active:
            branches = Branch.objects.defer(
                "country",
                "province",
                "city",
                "district",
                "date_created",
                "last_modified",
            ).filter(merchant__url_hash=url_hash)
        form = CreateCampaign(request.POST)
        if request.method == "POST":
            if form.is_valid():
                items = form.cleaned_data.pop("branch")
                cost = form.cleaned_data.pop("cost")
                group_id = uuid.uuid4()
                number_selected = len(items)
                cost = int(cost) / number_selected
                for item in items:
                    Campaign.objects.create(
                        branch=item,
                        cost=str(int(cost)),
                        group_id=group_id,
                        **form.cleaned_data,
                    )
                messages.success(request, "کمپین با موفقیت ساخته شد")
                return render(
                    request,
                    "create-campaign.html",
                    {"form": form, "branches": branches},
                )
            else:
                messages.error(request, f"{form.errors}")
        return render(
            request, "create-campaign.html", {"form": form, "branches": branches}
        )
    return render(request, "401.html", status=401)


@login_required
def edit_campaign(request, url_hash, group_id):
    """Using this, users can create campaigns."""
    if not perm_to_open(request, url_hash):
        return render(request, "401.html", status=401)
    if not request.user.profile.is_manager:
        return render(request, "401.html", status=401)
    campaign = (
        Campaign.objects.filter(group_id=group_id)
        .values("group_id")
        .annotate(
            campaign_name=Min("name"),
            campaign_start_date=Min("start_date"),
            campaign_end_date=Max("end_date"),
            campaign_cost=Sum("cost"),
            campaign_type=Min("campaign_type"),
            campaign_last_modified=Max("last_modified"),
            campaign_group_id=Min("group_id"),
        )
        .order_by("campaign_last_modified")
        .first()
    )
    branch_pks = (
        Campaign.objects.filter(group_id=campaign["campaign_group_id"])
        .values_list("branch__pk", flat=True)
        .distinct()
    )
    campaign["branch_pks"] = branch_pks
    selected_branch_pks = [row for row in campaign["branch_pks"]]
    start_date = convert_gregorian_to_jalali(campaign["campaign_start_date"])
    end_date = convert_gregorian_to_jalali(campaign["campaign_end_date"])
    if request.user.profile.merchant.url_hash == url_hash:
        if request.user.profile.is_manager and request.user.is_active:
            branches = Branch.objects.defer(
                "country",
                "province",
                "city",
                "district",
                "date_created",
                "last_modified",
            ).filter(merchant__url_hash=url_hash)
        form = CreateCampaign(request.POST)
        if request.method == "POST":
            if form.is_valid():
                items = form.cleaned_data.pop("branch")
                cost = form.cleaned_data.pop("cost")
                new_group_id = uuid.uuid4()
                number_selected = len(items)
                cost = int(float(cost)) / number_selected
                for item in items:
                    Campaign.objects.create(
                        branch=item,
                        cost=str(int(cost)),
                        group_id=new_group_id,
                        **form.cleaned_data,
                    )
                Campaign.objects.filter(group_id=group_id).delete()
                return redirect("campaign", request.user.profile.merchant.url_hash)
            else:
                messages.error(request, "لطفا همه فیلد ها را پر کنید")
        return render(
            request,
            "edit-campaign.html",
            {
                "form": form,
                "branches": branches,
                "campaign": campaign,
                "selected_branch_pks": selected_branch_pks,
                "start_date": start_date,
                "end_date": end_date,
            },
        )
    return render(request, "401.html", status=401)


@login_required
def delete_campaign(request, url_hash, group_id):
    """This deletes each campaign by its id."""
    if not perm_to_open(request, url_hash):
        return render(request, "401.html", status=401)
    if not request.user.profile.is_manager:
        return render(request, "401.html", status=401)
    campaign = (
        Campaign.objects.filter(group_id=group_id)
        .values("group_id")
        .annotate(
            campaign_name=Min("name"),
            campaign_start_date=Min("start_date"),
            campaign_end_date=Max("end_date"),
            campaign_cost=Sum("cost"),
            campaign_type=Min("campaign_type"),
            campaign_last_modified=Max("last_modified"),
            campaign_group_id=Min("group_id"),
        )
        .order_by("campaign_last_modified")
        .first()
    )
    branch_names = (
        Campaign.objects.filter(group_id=campaign["campaign_group_id"])
        .values_list("branch__name", flat=True)
        .distinct()
    )
    campaign["branch_names"] = ", ".join(branch_names)
    jalali_start_date = convert_gregorian_to_jalali(campaign["campaign_start_date"])
    jalali_end_date = convert_gregorian_to_jalali(campaign["campaign_end_date"])
    if request.method == "POST":
        Campaign.objects.filter(group_id=group_id).delete()
        return redirect("campaign", request.user.profile.merchant.url_hash)
    return render(
        request,
        "delete-campaign.html",
        {
            "campaign": campaign,
            "jalali_start_date": jalali_start_date,
            "jalali_end_date": jalali_end_date,
        },
    )


@login_required
def upload_excel_file_invoice(request, url_hash):
    """This uploads an excel file to create invoice objects accordingly."""
    # Rights
    if not perm_to_open(request, url_hash):
        return render(request, "401.html", status=401)
    if not request.user.has_perm("dashboard.add_invoice"):
        return render(request, "401.html", status=401)
    if not request.user.has_perm("dashboard.delete_invoice"):
        return render(request, "401.html", status=401)
    try:
        branches = Branch.objects.defer(
            "country", "province", "city", "district", "date_created", "last_modified"
        ).filter(merchant__url_hash=url_hash)
    except Exception:
        messages.error(request, "مرچنتی با این مشخصات یافت نشد.")
        return render(request, "upload_excel_invoice.html", {"error": f"{e}"})
    allowed_branches = []
    for branch in branches:
        allowed_branches.append(branch.pk)
    if request.method == "POST":
        form = UploadInvoiceExcel(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES["excel_file"]
            try:
                wb = openpyxl.load_workbook(excel_file)
                sheet = wb.active
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    date, branch, total_amount, total_items, total_product = (
                        row[:2] + row[3:]
                    )
                    if isinstance(date, str):
                        if date:
                            first_two = date[:2]
                            if first_two == "13" or first_two == "14":
                                year, month, day = map(int, date.split("-"))
                                jalali_date = jdatetime.date(year, month, day)
                                date = jalali_date.togregorian()
                            else:
                                date = datetime.strptime(date, "%Y-%m-%d").date()
                    if (
                        branch is not None
                        and total_amount is not None
                        and total_items is not None
                        and total_product is not None
                    ):
                        if int(branch) not in allowed_branches:
                            messages.warning(
                                request,
                                "یک یا چند ردیف دارای کد شعبی هستند که برای شما تعریف نشده است.",
                            )
                            return render(
                                request,
                                "upload_excel_invoice.html",
                                {
                                    "error": "از کد های زیر برای شعبه استفاده نمایید:",
                                    "branches": branches,
                                },
                            )
                        branch = branches.get(pk=int(branch))
                        Invoice.objects.update_or_create(
                            date=date,
                            branch=branch,
                            defaults={
                                "total_amount": int(total_amount),
                                "total_items": int(total_items),
                                "total_product": int(total_product),
                            },
                        )
                messages.success(request, "اطلاعات فروش با موفقیت آپلود شدند")
                return redirect(reverse("upload_excel_file_invoice", args=[url_hash]))
            except Exception as e:
                messages.error(
                    request,
                    "فایل مورد نظر یافت نشد. فرمت فایل و نام فایل آپلود شده را دوباره بررسی فرمایید. فایل مورد نظر باید طبق نمونه تمپلیت آپلود شود.",
                )
                return render(request, "upload_excel_invoice.html", {"error": f"{e}"})
        else:
            messages.error(request, "فایل مورد نظر باید دارای فرمت اکسل باشد.")
            return render(
                request, "upload_excel_invoice.html", {"error": f"{form.errors}"}
            )
    else:
        form = UploadInvoiceExcel()
        return render(request, "upload_excel_invoice.html", {"form": form})


@login_required
def invoices(request, url_hash):
    """This shows a list of all invoices. Filtering is also enabled."""
    # Rights
    if not perm_to_open(request, url_hash):
        return render(request, "401.html", status=401)
    if not request.user.has_perm("dashboard.add_invoice"):
        return render(request, "401.html", status=401)
    branches = Branch.objects.defer(
        "country", "province", "city", "district", "date_created", "last_modified"
    ).filter(merchant__url_hash=url_hash)
    permissions = (
        PermissionToViewBranch.objects.defer("date_created", "last_modified")
        .select_related("branch")
        .filter(user__pk=request.user.profile.pk)
    )
    permission_branch_keys = []
    for permission in permissions:
        permission_branch_keys.append(permission.branch.pk)
    allowed_branches = []
    if request.user.profile.is_manager:
        for branch in branches:
            allowed_branches.append(branch.pk)
    if not request.user.profile.is_manager and request.user.has_perm(
        "dashboard.add_invoice"
    ):
        allowed_branches = permission_branch_keys

    invoices = []
    branches = branches.filter(pk__in=allowed_branches)
    start_date_str = str(jalali_to_gregorian(request.GET.get("start-date")))
    end_date_str = str(jalali_to_gregorian(request.GET.get("end-date")))
    selected_branch_str = request.GET.getlist("branch")
    selected_branch = [int(item) for item in selected_branch_str]
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        if start_date_str and end_date_str:
            try:
                start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
                end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
                if len(selected_branch) == 0:
                    invoices = (
                        Invoice.objects.filter(branch__pk__in=allowed_branches)
                        .filter(date__range=(start_date, end_date))
                        .annotate(branch_name=F("branch__name"))
                    )
                else:
                    invoices = (
                        Invoice.objects.filter(branch__pk__in=selected_branch)
                        .filter(date__range=(start_date, end_date))
                        .annotate(branch_name=F("branch__name"))
                    )
            except Exception as e:
                print(e)
            return JsonResponse({"invoices": list(invoices.values())})
    return render(
        request, "invoices.html", {"invoices": invoices, "branches": branches}
    )


@login_required
def invoice_detail(request, invoice_pk):
    """This shows the detail of each invoice record. You can also delete and edit the total amount as well as the item numbers."""
    invoice = (
        Invoice.objects.defer("date_created", "last_modified", "branch__date_created")
        .select_related("branch")
        .get(pk=invoice_pk)
    )
    permissions = (
        PermissionToViewBranch.objects.defer("date_created", "last_modified")
        .select_related("branch")
        .filter(user__pk=request.user.profile.pk)
    )
    allowed_branches = []
    for permission in permissions:
        allowed_branches.append(permission.branch.pk)
    if not request.user.profile.is_manager:
        if invoice.branch.pk not in allowed_branches:
            return render(request, "401.html", status=401)
    form = InvoiceForm(instance=invoice)
    if request.method == "POST":
        form = InvoiceForm(request.POST, instance=invoice)
        if form.is_valid():
            form.save()
            messages.success(request, "تغییرات با موفقیت اعمال شد")
            return redirect(reverse("invoice_detail", args=[invoice_pk]))
        else:
            messages.error(request, f"{form.errors}")
            return redirect(reverse("invoice_detail", args=[invoice_pk]))
    return render(request, "invoice_detail.html", {"invoice": invoice, "form": form})


@login_required
def invoice_delete(request, invoice_pk):
    """This deletes the said invoice."""
    invoice = (
        Invoice.objects.defer("date_created", "last_modified")
        .select_related("branch")
        .get(pk=invoice_pk)
    )
    permissions = (
        PermissionToViewBranch.objects.defer("date_created", "last_modified")
        .select_related("branch")
        .filter(user__pk=request.user.profile.pk)
    )
    url_hash = request.user.profile.merchant.url_hash
    allowed_branches = []
    for permission in permissions:
        allowed_branches.append(permission.branch.pk)
    if not request.user.profile.is_manager:
        if invoice.branch.pk not in allowed_branches:
            return render(request, "401.html", status=401)
    invoice.delete()
    return redirect(reverse("invoices", args=[url_hash]))


def get_dates(start_date_str: str, end_date_str: str):
    """This creates a Python list within the Jalali specificed time period."""
    date_list = []
    start_date = (
        jdatetime.datetime.strptime(start_date_str, "%Y-%m-%d").date().togregorian()
    )
    end_date = (
        jdatetime.datetime.strptime(end_date_str, "%Y-%m-%d").date().togregorian()
    )
    current_date = start_date
    while current_date <= end_date:
        date_list.append(current_date)
        current_date += timedelta(days=1)
    date_list = [
        jdatetime.datetime.fromgregorian(date=date).date().strftime("%Y-%m-%d")
        for date in date_list
    ]
    return date_list


def create_excel_template(start_date: str, end_date: str, branch_list: dict):
    """It creates an excel file and styles it using Jalali start and end dates and a dictionary of numbers as keys and strings as values."""
    wb = openpyxl.Workbook()
    colors = [
        "FFF4CCCC",
        "FFFCE5CD",
        "FFFFF2CC",
        "FFD9EAD3",
        "FFD0E0E3",
        "FFCFE2F3",
        "FFD9D2E9",
        "FFEAD1DC",
        "FFEEEEEE",
        "FFBCBCBC",
    ]
    color_mapper = {}
    for item in branch_list.keys():
        color_mapper[item] = random.sample(colors, 1)[0]
    ws = wb.active
    ws.title = "تمپلیت آپلود اطلاعات فروش شعب"
    ws.append(
        ["تاریخ", "کد شعبه", "نام شعبه", "مبلغ فاکتور", "تعداد فاکتور", "تعداد کالا"]
    )
    branches = branch_list
    dates = get_dates(start_date, end_date)
    for branch_id, branch_name in branches.items():
        for date in dates:
            ws.append([date, branch_id, branch_name, "", ""])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[1].value in color_mapper.keys():
            for cell in row:
                cell.fill = openpyxl.styles.PatternFill(
                    start_color=color_mapper[row[1].value],
                    end_color=color_mapper[row[1].value],
                    fill_type="solid",
                )
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@login_required
def excel_template_generator(request, url_hash):
    """This creates an excel template file for uploading to the system without any errors."""
    # Rights
    if not perm_to_open(request, url_hash):
        return render(request, "401.html", status=401)
    permissions = (
        PermissionToViewBranch.objects.defer("date_created", "last_modified")
        .select_related("branch")
        .filter(user__pk=request.user.profile.pk)
    )
    allowed_branches = []
    if not request.user.profile.is_manager:
        for permission in permissions:
            allowed_branches.append(permission.branch.pk)
        branches = (
            Branch.objects.defer(
                "country",
                "province",
                "city",
                "district",
                "date_created",
                "last_modified",
            )
            .filter(merchant__url_hash=url_hash)
            .filter(pk__in=allowed_branches)
        )
    else:
        branches = Branch.objects.defer(
            "country", "province", "city", "district", "date_created", "last_modified"
        ).filter(merchant__url_hash=url_hash)
    try:
        start_date_str = str(request.GET.get("start-date"))
        end_date_str = str(request.GET.get("end-date"))
        branches_str = request.GET.getlist("branch")
        branches_int = [int(i) for i in branches_str]
        branch_dict = {}
        for branch in branches:
            if branch.pk in branches_int:
                branch_dict[branch.pk] = branch.name
        if start_date_str and end_date_str:
            excel_file = create_excel_template(
                start_date_str, end_date_str, branch_dict
            )
            response = HttpResponse(
                excel_file,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="template.xlsx"'
            return response
    except Exception as e:
        print(e)
    return render(request, "create-excel-template.html", {"branches": branches})


@login_required
def invoice_counter(request, url_hash):
    # Rights
    if not perm_to_open(request, url_hash):
        return render(request, "401.html", status=401)
    # Permissions
    allowed_branches = []
    campaign_list = []
    profile = request.user.profile
    all_branches = Branch.objects.defer(
        "country", "province", "city", "district", "date_created", "last_modified"
    ).filter(merchant__url_hash=url_hash)
    if not profile.is_manager:
        permissions = (
            PermissionToViewBranch.objects.defer("date_created", "last_modified")
            .select_related("branch")
            .filter(user__pk=profile.pk)
        )
        for permission in permissions:
            allowed_branches.append(permission.branch.pk)
    else:
        for item in all_branches:
            allowed_branches.append(item.pk)
    try:
        start_date_str = str(jalali_to_gregorian(request.GET.get("start-date")))
        end_date_str = str(jalali_to_gregorian(request.GET.get("end-date")))
        branches_str = request.GET.getlist("branch")
    except Exception as e:
        print(e)
    branches = []
    dates = []
    total_amount = []
    total_items = []
    total_products = []
    invoices = (
        Invoice.objects.defer("date_created", "last_modified")
        .select_related("branch", "branch__merchant")
        .filter(branch__pk__in=allowed_branches)
        .values("date")
        .annotate(
            sum_total_amount=Sum("total_amount"),
            sum_total_items=Sum("total_items"),
            sum_total_products=Sum("total_product"),
        )
        .order_by("date")
    )
    if start_date_str != "None" and end_date_str != "None":
        try:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
            invoices = invoices.filter(date__range=(start_date, end_date))
            campaigns = (
                Campaign.objects.filter(
                    branch__merchant__url_hash=request.user.profile.merchant.url_hash
                )
                .filter(branch__pk__in=allowed_branches)
                .filter(Q(start_date__lte=end_date) & Q(end_date__gte=start_date))
                .values("group_id")
                .annotate(
                    campaign_name=Min("name"),
                    campaign_start_date=Min("start_date"),
                    campaign_end_date=Max("end_date"),
                    campaign_cost=Sum("cost"),
                    campaign_type=Min("campaign_type"),
                    campaign_last_modified=Max("last_modified"),
                    campaign_group_id=Min("group_id"),
                )
                .order_by("-campaign_last_modified")
            )
            campaign_list = []
            dictionary = {}
            for campaign in campaigns:
                dictionary = {
                    "campaign_name": campaign["campaign_name"],
                    "campaign_start_date": campaign["campaign_start_date"],
                    "campaign_end_date": campaign["campaign_end_date"],
                    "campaign_cost": campaign["campaign_cost"],
                    "campaign_type": campaign["campaign_type"],
                }
                campaign_list.append(dictionary)
        except Exception as e:
            print(e)
        if branches_str:
            try:
                branches = [
                    int(branch_str)
                    for branch_str in branches_str
                    if int(branch_str) in allowed_branches
                ]
            except ValueError:
                branches = []
            if branches:
                invoices = invoices.filter(branch__pk__in=branches)
                campaigns = (
                    Campaign.objects.filter(
                        branch__merchant__url_hash=request.user.profile.merchant.url_hash
                    )
                    .filter(branch__pk__in=allowed_branches)
                    .filter(branch__pk__in=branches)
                    .filter(Q(start_date__lte=end_date) & Q(end_date__gte=start_date))
                    .values("group_id")
                    .annotate(
                        campaign_name=Min("name"),
                        campaign_start_date=Min("start_date"),
                        campaign_end_date=Max("end_date"),
                        campaign_cost=Sum("cost"),
                        campaign_type=Min("campaign_type"),
                        campaign_last_modified=Max("last_modified"),
                        campaign_group_id=Min("group_id"),
                    )
                    .order_by("-campaign_last_modified")
                )
        dates = [str(row["date"].strftime("%Y-%m-%d")) for row in invoices]
        total_items = [float(row["sum_total_items"]) for row in invoices]
        total_amount = [float(row["sum_total_amount"] // 10000000) for row in invoices]
        total_products = [float(row["sum_total_products"]) for row in invoices]
        campaign_list = []
        dictionary = {}
        for campaign in campaigns:
            dictionary = {
                "campaign_name": campaign["campaign_name"],
                "campaign_start_date": campaign["campaign_start_date"],
                "campaign_end_date": campaign["campaign_end_date"],
                "campaign_cost": campaign["campaign_cost"],
                "campaign_type": campaign["campaign_type"],
            }
            campaign_list.append(dictionary)

        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            print(connection.queries, len(connection.queries))
            return JsonResponse(
                {
                    "dates": dates,
                    "total_items": total_items,
                    "total_amount": total_amount,
                    "campaigns": campaign_list,
                    "total_products": total_products,
                }
            )
    if not profile.is_manager:
        all_branches = all_branches.filter(pk__in=allowed_branches)
    print(connection.queries, len(connection.queries))
    return render(
        request,
        "invoice-counter.html",
        {
            "invoices": invoices,
            "branches": all_branches,
            "dates": json.dumps(dates),
            "total_amount": json.dumps(total_amount),
            "total_items": json.dumps(total_items),
            "total_products": json.dumps(total_products),
        },
    )


@login_required
def analysis(request, url_hash):
    if not perm_to_open(request, url_hash):
        return render(request, "401.html", status=401)
    if not request.user.profile.is_manager:
        return render(request, "401.html", status=401)
    allowed_branches = []
    campaign_list = []
    all_branches = Branch.objects.defer(
        "country", "province", "city", "district", "date_created", "last_modified"
    ).filter(merchant__url_hash=url_hash)

    for item in all_branches:
        allowed_branches.append(item.pk)
    print(allowed_branches)
    try:
        start_date_str = str(jalali_to_gregorian(request.GET.get("start-date")))
        end_date_str = str(jalali_to_gregorian(request.GET.get("end-date")))
        branches_str = request.GET.getlist("branch")
    except Exception as e:
        print(e)
    branches = []
    queryset_no_branch_filter = []
    invoices = []
    queryset = []
    campaigns = (
        Campaign.objects.defer("date_created", "last_modified")
        .select_related("branch")
        .filter(branch__merchant__url_hash=url_hash, branch__pk__in=allowed_branches)
    )
    if start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
            invoices = (
                Invoice.objects.defer("date_created", "last_modified")
                .filter(
                    branch__pk__in=allowed_branches, date__range=(start_date, end_date)
                )
                .values("date")
                .annotate(
                    sum_total_amount=Sum("total_amount"),
                    sum_total_items=Sum("total_items"),
                    sum_total_products=Sum("total_product"),
                )
                .order_by("date")
            )
            queryset = (
                PeopleCounting.objects.filter(
                    merchant__url_hash=url_hash,
                    branch__pk__in=allowed_branches,
                    date__range=(start_date, end_date),
                )
                .values("date")
                .annotate(total_entry=Sum("entry"), total_exit=Sum("exit"))
                .order_by("date")
            )
            queryset_no_branch_filter = queryset.filter(
                date__range=(start_date, end_date)
            )
            campaigns = campaigns.filter(
                Q(start_date__lte=end_date) & Q(end_date__gte=start_date)
            )
            campaign_list = []
            for campaign in campaigns:
                dictionary = {
                    "campaign_name": campaign.name,
                    "campaign_start_date": campaign.start_date,
                    "campaign_end_date": campaign.end_date,
                    "campaign_branch_name": campaign.branch.name,
                    "campaign_type": campaign.campaign_type,
                    "campaign_cost": campaign.cost,
                }
                campaign_list.append(dictionary)
        except Exception as e:
            print(e)
        if branches_str:
            try:
                branches = [
                    int(branch_str)
                    for branch_str in branches_str
                    if int(branch_str) in allowed_branches
                ]
            except ValueError:
                branches = []
            if branches:
                invoices = invoices.filter(branch__pk__in=branches)
                queryset = queryset.filter(branch__pk__in=branches)
                campaigns = campaigns.filter(branch__pk__in=branches)

        dates_queryset = [str(row["date"].strftime("%Y-%m-%d")) for row in queryset]
        dates_invoices = [str(row["date"].strftime("%Y-%m-%d")) for row in invoices]
        total_items = [float(row["sum_total_items"]) for row in invoices]
        total_amount = [float(row["sum_total_amount"]) for row in invoices]
        total_products = [row["sum_total_products"] for row in invoices]
        total_products_avg = [
            round(float(a / b if b != 0 else 0), 1)
            for a, b in zip(total_products, total_items)
        ]
        entry_totals = [float(row["total_entry"]) for row in queryset]
        entry_overalls = [
            float(row["total_entry"]) for row in queryset_no_branch_filter
        ]
        campaign_list = []
        for campaign in campaigns:
            dictionary = {
                "campaign_name": campaign.name,
                "campaign_start_date": campaign.start_date,
                "campaign_end_date": campaign.end_date,
                "campaign_branch_name": campaign.branch.name,
                "campaign_type": campaign.campaign_type,
                "campaign_cost": campaign.cost,
            }
            campaign_list.append(dictionary)
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            print(connection.queries, len(connection.queries))
            return JsonResponse(
                {
                    "dates_queryset": dates_queryset,
                    "dates_invoices": dates_invoices,
                    "entry_totals": entry_totals,
                    "total_items": total_items,
                    "total_products_avg": total_products_avg,
                    "total_amount": total_amount,
                    "entry_overalls": entry_overalls,
                    "campaigns": campaign_list,
                }
            )
    print(connection.queries, len(connection.queries))
    return render(
        request,
        "analysis.html",
        {
            "branches": all_branches,
            "dates_queryset": json.dumps(dates_queryset),
            "dates_invoices": json.dumps(dates_invoices),
            "entry_totals": json.dumps(entry_totals),
            "total_amount": json.dumps(total_amount),
            "total_items": json.dumps(total_items),
            "entry_overalls": json.dumps(entry_overalls),
            "total_products_avg": json.dumps(total_products_avg),
        },
    )


@login_required
def campaign_detail(request, campaign_id):
    # Permissions
    allowed_branches = []
    profile = request.user.profile
    all_branches = Branch.objects.defer(
        "country", "province", "city", "district", "date_created", "last_modified"
    ).filter(merchant__url_hash=profile.merchant.url_hash)
    if not profile.is_manager:
        permissions = (
            PermissionToViewBranch.objects.defer("date_created", "last_modified")
            .select_related("branch")
            .filter(user__pk=profile.pk)
        )
        for permission in permissions:
            allowed_branches.append(permission.branch.pk)
    else:
        for item in all_branches:
            allowed_branches.append(item.pk)
    campaign = Campaign.objects.get(pk=campaign_id)
    if campaign.branch.pk not in allowed_branches:
        return render(request, "401.html", status=401)
    people_counting_info = PeopleCounting.objects.filter(
        branch=campaign.branch, date__range=(campaign.start_date, campaign.end_date)
    )
    invoice_info = Invoice.objects.filter(
        branch=campaign.branch, date__range=(campaign.start_date, campaign.end_date)
    )
    people_counting_series = [float(row.entry) for row in people_counting_info]
    invoice_sales_series = [float(row.total_amount) for row in invoice_info]
    invoice_count_series = [float(row.total_items) for row in invoice_info]
    dates = [str(row.date.strftime("%Y-%m-%d")) for row in people_counting_info]
    return render(
        request,
        "campaign-solo-analysis.html",
        {
            "campaign": campaign,
            "dates": json.dumps(dates),
            "entry_totals": json.dumps(people_counting_series),
            "total_amount": json.dumps(invoice_sales_series),
            "total_items": json.dumps(invoice_count_series),
        },
    )


@login_required
def campaign_comparison(request, url_hash):
    # Rights
    if not perm_to_open(request, url_hash):
        return render(request, "401.html", status=401)
    if not request.user.profile.is_manager:
        return render(request, "401.html", status=401)
    grouped_campaigns = []
    campaigns = []
    branches = Branch.objects.filter(merchant__url_hash=url_hash)
    try:
        selected_branch_str = request.GET.getlist("branch")
        names = request.GET.getlist("name")
        selected_branch = [int(item) for item in selected_branch_str]
    except Exception as e:
        print(e)
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        if names:
            try:
                grouped_campaigns = (
                    Campaign.objects.defer("last_modified", "date_created")
                    .filter(
                        branch__merchant__url_hash=request.user.profile.merchant.url_hash,
                    )
                    .filter(name__in=names)
                    .values("group_id")
                    .annotate(
                        campaign_name=Min("name"),
                        campaign_start_date=Min("start_date"),
                        campaign_end_date=Max("end_date"),
                        campaign_cost=Sum("cost"),
                        campaign_type=Min("campaign_type"),
                        campaign_last_modified=Max("last_modified"),
                        campaign_group_id=Min("group_id"),
                    )
                    .order_by("-campaign_last_modified")
                )
                campaigns = (
                    Campaign.objects.defer("last_modified", "date_created")
                    .filter(
                        branch__merchant__url_hash=request.user.profile.merchant.url_hash
                    )
                    .filter(name__in=names)
                    .annotate(branch_name=F("branch__name"))
                )
                if grouped_campaigns:
                    for campaign in grouped_campaigns:
                        branch_names = (
                            Campaign.objects.filter(
                                group_id=campaign["campaign_group_id"]
                            )
                            .values_list("branch__name", flat=True)
                            .distinct()
                        )
                        campaign["branch_names"] = ", ".join(branch_names)
            except Exception as e:
                print(e)
            return JsonResponse(
                {
                    "grouped_campaigns": list(grouped_campaigns),
                    "campaigns": list(campaigns.values()),
                }
            )
    return render(
        request,
        "campaign-comparison.html",
        {
            "grouped_campaigns": grouped_campaigns,
            "campaigns": campaigns,
            "branches": branches,
        },
    )


def ping_ip(ip, timeout=15):
    try:
        result = subprocess.run(
            ["ping", "-c", "1", "-W", str(timeout), ip],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            timeout=timeout + 2,
        )
        return result.returncode == 0
    except subprocess.TimeoutExpired:
        print(f"Timeout: Ping to {ip} took too long.")
        return False
    except Exception as e:
        print(f"Error: {e}")
        return False


@login_required
def camera_list(request, url_hash):
    # Rights
    if not perm_to_open(request, url_hash):
        return render(request, "401.html", status=401)
    if not request.user.profile.is_manager:
        return render(request, "401.html", status=401)
    cameras = Cam.objects.filter(
        merchant__url_hash=request.user.profile.merchant.url_hash
    )
    for camera in cameras:
        camera.status = ping_ip(camera.ip)
        camera.save()
    return render(request, "camera-list.html", {"cameras": cameras})


def holiday_spotter(request, year, month, day):
    response = requests.get(f"https://holidayapi.ir/jalali/{year}/{month}/{day}")
    if response.status_code == 200:
        data = response.json()
        return JsonResponse(data)
    return JsonResponse({"error": "no response"}, status=response.status_code)


@login_required
def alert_menu(request, url_hash):
    # Rights
    if not perm_to_open(request, url_hash):
        return render(request, "401.html", status=401)
    return render(request, "alert-menu.html")


@login_required
def alert_form_sms(request, url_hash):
    # Rights
    if not perm_to_open(request, url_hash):
        return render(request, "401.html", status=401)
    form = AlertCameraMalfunctionForm()
    if request.method == "POST":
        form = AlertCameraMalfunctionForm(request.POST)
        if form.is_valid():
            contact = form.save(commit=False)
            contact.merchant = request.user.profile.merchant
            contact = form.save()
            messages.success(request, "مخاطب با موفقیت ساخته شد")
        else:
            messages.error(request, f"{form.errors}")
            return render(
                request,
                "alert-form-sms.html",
                {"form": form, "error": f"{form.errors}"},
            )
    return render(request, "alert-form-sms.html", {"form": form})


@login_required
def alert_from_sms_contact_list(request, url_hash):
    # Rights
    if not perm_to_open(request, url_hash):
        return render(request, "401.html", status=401)
    contacts = AlertCameraMalfunction.objects.filter(
        merchant__url_hash=url_hash
    ).order_by("-pk")
    return render(request, "alert_from_sms_contact_list.html", {"contacts": contacts})


@login_required
def alert_from_sms_contact_list_detail(request, contact_id):
    if not request.user.profile.mobile:
        return render(request, "401.html", status=401)
    contact = get_object_or_404(AlertCameraMalfunction, pk=contact_id)
    messages = AlertCameraMalfunctionMessage.objects.filter(contact=contact).order_by(
        "-pk"
    )
    return render(
        request, "alert_from_sms_contact_list_detail.html", {"messages": messages}
    )


@login_required
def alert_form_sms_edit(request, contact_id):
    if not request.user.profile.mobile:
        return render(request, "401.html", status=401)
    contact = get_object_or_404(AlertCameraMalfunction, pk=contact_id)
    form = AlertCameraMalfunctionForm(instance=contact)
    if request.method == "POST":
        form = AlertCameraMalfunctionForm(request.POST, instance=contact)
        if form.is_valid():
            edit_contact = form.save(commit=False)
            edit_contact.merchant = request.user.profile.merchant
            edit_contact.save()
            messages.success(request, "مخاطب با موفقیت تغییر یافت")
    else:
        messages.error(request, f"{form.errors}")
        return render(
            request,
            "alert_form_sms_edit.html",
            {"contact": contact, "form": form, "error": f"{form.errors}"},
        )
    return render(
        request, "alert_form_sms_edit.html", {"form": form, "contact": contact}
    )


def alert_form_social(request, url_hash):
    # Rights
    if not perm_to_open(request, url_hash):
        return render(request, "401.html", status=401)
    return render(request, "alert-form-social.html")


def grouped_campaign_search_as_type(request):
    q = request.GET.get("q", "").strip()
    if q == "":
        return JsonResponse([], safe=False)
    grouped_campaigns = (
        Campaign.objects.defer("last_modified", "date_created")
        .filter(
            branch__merchant__url_hash=request.user.profile.merchant.url_hash,
            name__icontains=q
        )
        .values("group_id")
        .annotate(
            campaign_name=Min("name"),
            campaign_start_date=Min("start_date"),
            campaign_end_date=Max("end_date"),
            campaign_cost=Sum("cost"),
            campaign_type=Min("campaign_type"),
            campaign_last_modified=Max("last_modified"),
            campaign_group_id=Min("group_id"),
        )
        .order_by("-campaign_last_modified")
    )
    names = list(grouped_campaigns.values("name"))
    loop_index = 0
    for item in names:
        loop_index += 1
        name_value = item.pop("name")
        item["id"] = loop_index
        item["text"] = name_value
    print(names)
    return JsonResponse(names, safe=False)


def single_campaign_search_as_type(request):
    q = request.GET.get("q", "").strip()
    if q == "":
        return JsonResponse([], safe=False)
    campaigns = Campaign.objects.filter(branch__merchant__url_hash=request.user.profile.merchant.url_hash, name__icontains=q).order_by("-pk")[:10]
    names_with_branches = list(campaigns.values("name", "branch__name"))
    print(names_with_branches)
    # [{'name': 'تست 15', 'branch__name': 'اقدسیه'}]
    loop_index = 0
    for item in names_with_branches:
        campaign_name = item.pop("name")
        branch_name = item.pop("branch__name")
        text = f"{campaign_name} - {branch_name}"
        loop_index += 1
        item["id"] = loop_index
        item["text"] = text
    print(names_with_branches)
    return JsonResponse(names_with_branches, safe=False)

