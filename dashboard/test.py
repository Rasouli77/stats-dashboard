import jdatetime
import datetime
import openpyxl
import random

import openpyxl.styles
def get_dates(start_date_str: str, end_date_str:str):
    date_list = []
    start_date = jdatetime.datetime.strptime(start_date_str, "%Y-%m-%d").date().togregorian()
    end_date = jdatetime.datetime.strptime(end_date_str, "%Y-%m-%d").date().togregorian()
    current_date = start_date
    while current_date <= end_date:
        date_list.append(current_date)
        current_date += datetime.timedelta(days=1)
    date_list = [jdatetime.datetime.fromgregorian(date=date).date().strftime("%Y-%m-%d") for date in date_list]

    return date_list

x = get_dates("1404-05-01", "1404-05-10")
print(x)

def create_excel_template(start_date: str, end_date: str, branch_list: list, filename="x.xlsx"):
    wb = openpyxl.Workbook()
    colors = [
    "FFFFF2CC", "FFFCE5CD", "FFEAD1DC", "FFD9EAD3", "FFD0E0E3",
    "FFC9DAF8", "FFF4CCCC", "FFD9D2E9", "FFB6D7A8", "FFCFE2F3",
    "FFF6B26B", "FFA2C4C9", "FFE6B8AF", "FFB4A7D6", "FFFFD966",
    "FFB7E1CD", "FFF9CB9C", "FFB6DDE8", "FFD5A6BD", "FFA4C2F4",
    "FFFFF9C4", "FFDCEDC8", "FFFFECB3", "FFE1BEE7", "FFC8E6C9",
    "FFFFCDD2", "FFBBDEFB", "FFD1C4E9", "FFFFCCBC", "FFB3E5FC",
    "FFFFEBEE", "FFE3F2FD", "FFEDE7F6", "FFF3E5F5", "FFFFF3E0",
    "FFF1F8E9", "FFFBE9E7", "FFE0F2F1", "FFF0F4C3", "FFF8BBD0",
    "FFC5CAE9", "FFFFCCFF", "FFE6EE9C", "FFFFAB91", "FFB2EBF2",
    "FFD7CCC8", "FFDCEDC8", "FFF5F5DC", "FFE0F7FA", "FFF0FFF0"
    ]
    color_mapper = {}
    for item in branch_list:
        color_mapper[item] = random.sample(colors, 1)[0]
    print(color_mapper)
    ws = wb.active
    ws.title = "تمپلیت آپلود اطلاعات فروش شعب"
    ws.append(["تاریخ", " شعبه", "مبلغ فاکتور", "تعداد فاکتور"])
    branches = branch_list
    dates = get_dates(start_date, end_date)
    for branch in branches:
        for date in dates:
            ws.append([date, branch, "", ""])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[1].value in color_mapper.keys():
            for cell in row:
                print(color_mapper[row[1].value])
                cell.fill = openpyxl.styles.PatternFill(start_color=color_mapper[row[1].value], end_color=color_mapper[row[1].value], fill_type="solid")
    wb.save(filename)

create_excel_template("1404-05-01", "1404-05-10", [1, 2, 3, 4])